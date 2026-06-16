"""21 - Extract & clean AHA's gold set from the monthly Excel workbook.

Reads the verification workbook (one tab per month), normalizes DOIs, classifies
each row into a gold class from the Status column (+ N/A note parsing), and pulls
out AHA-format grant IDs.

Writes:
  data/goldset_full.csv  (GITIGNORED — includes notes/PI, for our own diagnosis)
  data/goldset.csv       (checked in — de-identified: no PI, no verbatim notes)

Run:  PYTHONIOENCODING=utf-8 py -3 21_extract_goldset.py
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config as C

DOI_PREFIXES = ("https://doi.org/", "http://doi.org/", "doi.org/", "doi:")


def norm_doi(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    for p in DOI_PREFIXES:
        if s.startswith(p):
            s = s[len(p):]
    s = s.strip().rstrip(".,;:")
    return s if s.startswith("10.") else ""


def aha_grants(raw) -> list:
    if not raw:
        return []
    out = []
    for tok in str(raw).replace(",", ";").split(";"):
        t = tok.strip()
        if C.AHA_GRANT_RE.match(t):
            out.append(t)
    return out


def gold_class(status: str, na_cat: str) -> str:
    """Map verification -> gold class for matching.
      positive          : verified AHA-funded (should link to AHA funder)
      positive_no_grant : theirs but no grant ID cited (funder-yes, grant-no)
      negative          : verified NOT AHA-funded (should NOT link)
      ambiguous         : N/A without a clear note
      pending           : not yet verified / in progress
    """
    if status in C.POSITIVE_STATUS:
        return "positive"
    if status == "N/A":
        if na_cat in ("remove_erroneous", "not_funded_disclosure",
                      "aha_authored", "doc_type_excluded"):
            return "negative"
        if na_cat == "theirs_no_grant_id":
            return "positive_no_grant"
        return "ambiguous"
    return "pending"   # Pending / Non-responsive / blank


def main() -> None:
    wb = openpyxl.load_workbook(C.RAW_XLSX, data_only=True)
    rows = []
    for sheet in C.MONTH_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  (skip missing sheet {sheet!r})")
            continue
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            g = lambda k: ws.cell(r, C.COL[k]).value
            doi = norm_doi(g("doi"))
            pmid = g("pmid")
            if not doi and not pmid:        # skip summary / blank rows
                continue
            status = str(g("status")).strip() if g("status") is not None else ""
            note = g("notes")
            na_cat = C.na_note_category(note) if status == "N/A" else ""
            grants = aha_grants(g("grants"))
            pi = " ".join(str(g(k)).strip() for k in ("pi_first", "pi_last")
                          if g(k)).strip()
            rows.append({
                "month": sheet,
                "doi": doi,
                "pmid": str(pmid).strip() if pmid else "",
                "pmcid": str(g("pmcid")).strip() if g("pmcid") else "",
                "status": status,
                "na_note_category": na_cat,
                "gold_class": gold_class(status, na_cat),
                "aha_grant_ids": "|".join(grants),
                "n_aha_grants": len(grants),
                "funders_raw": str(g("funder")).strip() if g("funder") else "",
                "pmcid_status": str(g("pmcid_status")).strip() if g("pmcid_status") else "",
                # PII / verbatim — full file only
                "_pi_name": pi,
                "_note": str(note).strip() if note else "",
            })

    df = pd.DataFrame(rows)
    # de-dup exact DOI repeats across tabs (keep first, note collisions)
    dupe = df[df["doi"] != ""].duplicated(subset="doi", keep=False).sum()

    df.to_csv(C.DATA / "goldset_full.csv", index=False, encoding="utf-8")
    public = df.drop(columns=["_pi_name", "_note"])
    public.to_csv(C.DATA / "goldset.csv", index=False, encoding="utf-8")

    print(f"rows: {len(df)}   (DOIs present: {(df.doi!='').sum()}, "
          f"duplicate-DOI rows: {dupe})")
    print("\ngold_class:\n", df.gold_class.value_counts().to_string())
    print("\nstatus:\n", df.status.replace('', '(blank)').value_counts().to_string())
    print("\nN/A note categories:\n",
          df[df.status == 'N/A'].na_note_category.value_counts().to_string())
    print(f"\nrows with >=1 AHA-format grant: {(df.n_aha_grants>0).sum()}")
    print("wrote data/goldset.csv (public) + data/goldset_full.csv (gitignored)")


if __name__ == "__main__":
    main()
