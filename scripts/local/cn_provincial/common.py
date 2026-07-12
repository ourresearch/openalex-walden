#!/usr/bin/env python3
"""Shared machinery for Chinese provincial S&T award-list scrapers.

See __init__.py for the design overview. The public entry point is
`run_province(cfg, argv)`, which a per-province runner calls with its
ProvinceConfig. It implements the full standard pipeline:

    1. Enumerate the listing column(s) via the Hanweb `dataproxy.jsp` proxy
       that every provincial "大汉版通" (Hanweb) CMS exposes.
    2. Filter article titles to award-list announcements (cfg.article_pattern).
    3. Fetch each article, download its attachments (xls/xlsx/pdf/docx/doc),
       or parse an inline HTML table.
    4. Header-driven table parsing: detect the header row by Chinese column
       keywords and map to the common output schema. This is what makes the
       framework reusable -- a new province usually needs no format code.
    5. Checkpoint after every article (resumable).
    6. Write a flat parquet (all string columns) and upload with the
       runbook 1.4 shrink-check.

The parquet is deliberately FLAT (one row per funded project, with
funder_award_id / display_name / lead_family_name / institution / ... columns).
The Databricks notebook assembles the §2.1 struct schema, exactly as
CreateNSFCAwards does. Chinese PI names follow the NSFC precedent: the full
name goes in family_name and given_name is left NULL (Chinese order is
family-first and unsplittable without a dictionary).
"""

from __future__ import annotations

# --- Windows UTF-8 compatibility shim (fleet-fix 2026-05-22) ---
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    sys.stderr.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
except (AttributeError, ValueError):
    pass

import builtins as _builtins_utf8
if sys.platform == "win32":
    import pathlib as _pathlib_utf8

    _orig_wt = _pathlib_utf8.Path.write_text
    def _wt(self, data, encoding=None, errors=None, newline=None):
        return _orig_wt(self, data, encoding=encoding or "utf-8", errors=errors, newline=newline)
    _pathlib_utf8.Path.write_text = _wt

    _orig_rt = _pathlib_utf8.Path.read_text
    def _rt(self, encoding=None, errors=None, newline=None):
        return _orig_rt(self, encoding=encoding or "utf-8", errors=errors, newline=newline)
    _pathlib_utf8.Path.read_text = _rt

    _orig_open = _builtins_utf8.open
    def _open_utf8(file, mode="r", buffering=-1, encoding=None, errors=None, newline=None, closefd=True, opener=None):
        if "b" not in mode and encoding is None:
            encoding = "utf-8"
        return _orig_open(file, mode, buffering, encoding, errors, newline, closefd, opener)
    _builtins_utf8.open = _open_utf8
# --- end shim ---

import argparse
import json
import re
import subprocess
import time
import warnings
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Optional
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests

warnings.filterwarnings("ignore", message="Workbook contains no default style")
requests.packages.urllib3.disable_warnings()  # provincial portals use self-signed / mismatched certs

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

S3_BUCKET = "openalex-ingest"


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@dataclass
class ListingColumn:
    """One Hanweb `dataproxy.jsp` column to enumerate.

    Every provincial "大汉版通" CMS serves its notice list from
    `{base}/module/web/jpage/dataproxy.jsp` with these POST params. Grab them
    once by watching the col page's network tab (or read the `<nextgroup>`
    CDATA in the col index HTML). `stride` is how many records the server
    returns per group (empirically ~45 for these portals); `perpage` MUST match
    the column's native page size or the proxy returns a `dataStore=['csrf']`
    stub instead of records.
    """
    referer: str                      # the col/colNNNNN/index.html page
    columnid: str
    unitid: str
    webid: str
    webname: str                      # the CMS site name (Chinese), url-decoded
    path: str = "/"                   # 'path' param; Shandong uses full base URL, Jiangsu uses '/'
    perpage: int = 15
    stride: int = 45
    needs_cookie_prime: bool = True   # GET the referer first to obtain a session cookie (Jiangsu requires this)

    def proxy_url(self, base: str) -> str:
        return urljoin(base, "/module/web/jpage/dataproxy.jsp")

    def post_data(self) -> dict:
        return {
            "col": "1", "appid": "1", "webid": self.webid, "path": self.path,
            "columnid": self.columnid, "sourceContentType": "1",
            "unitid": self.unitid, "webname": self.webname, "permissiontype": "0",
        }


@dataclass
class ProvinceConfig:
    provenance: str                   # e.g. "shandong_nsf"
    funder_id: int                    # OpenAlex numeric funder id (verified via API)
    funder_display_name: str
    base_url: str                     # e.g. "http://kjt.shandong.gov.cn"
    listing_columns: list[ListingColumn]
    # Select award-list announcements from listing titles. Both must be present:
    article_pattern: str              # regex the title MUST match (e.g. 自然科学基金 + 拟立项)
    article_exclude: str = r"$^"      # regex to drop (评审专家/答辩成绩/申报通知/指南)
    # Attachment link text/href hints. Default handles xls/xlsx/pdf/doc/docx.
    attachment_ext: str = r"\.(xlsx?|pdf|docx?|et|wps)([?\"]|$)|downfile"
    # If an article carries no attachment, try to parse the inline HTML <table>.
    allow_inline_table: bool = False
    default_country: str = "China"
    # Optional: derive a funder_scheme from the article title when the table lacks one.
    scheme_from_title: Optional[Callable[[str], Optional[str]]] = None
    # Optional: derive start_year from article title (fallback when table lacks a year col).
    year_from_title: Optional[Callable[[str], Optional[int]]] = None
    request_pause: float = 0.8


# ---------------------------------------------------------------------------
# Column-keyword header detection (the reusable core)
# ---------------------------------------------------------------------------

# Order matters: more specific keys first. Each maps a normalized column role
# to the Chinese header substrings that identify it across provinces.
HEADER_KEYWORDS: list[tuple[str, list[str]]] = [
    ("funder_award_id", ["立项编号", "项目编号", "申报编号", "受理号", "批准号", "课题编号", "编号"]),
    ("institution",     ["依托单位", "申报单位", "承担单位", "推荐单位", "单位名称", "所在单位", "单位"]),
    ("lead_name",       ["项目负责人", "团队负责人", "申报者", "申报人", "负责人", "主持人", "姓名"]),
    ("display_name",    ["项目名称", "课题名称", "项目及方向", "项目名", "专项及项目名称", "名称"]),
    ("funder_scheme",   ["项目类别", "项目类型", "计划类别", "资助类别", "类别", "类型"]),
    ("amount",          ["资助金额", "资助经费", "经费", "金额", "拨款", "预算"]),
    ("start_year",      ["年度", "立项年度", "年份"]),
    ("raw_extra",       ["所属学科", "学科", "备注", "平均分", "答辩得分", "得分", "推荐意见", "建议情况"]),
]

# Cells that are section headers, not projects (e.g. "一 高速切削...", "1.区块链...").
SECTION_ROW = re.compile(r"^[一二三四五六七八九十]+\s|^\d+\s*[.、]\s*\D|^附件")
# A plausible PI name: 2-4 CJK chars, OR a latin name (foreign PIs appear in SD 2020).
CJK = re.compile(r"[一-鿿]")


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def detect_header(rows: list[list[Any]], scan: int = 6) -> Optional[tuple[int, dict[str, int]]]:
    """Find the header row and map column roles -> column index.

    Returns (header_row_index, {role: col_idx}) or None if no title column is
    found in the first `scan` rows.
    """
    best = None
    for ri in range(min(scan, len(rows))):
        # Strip ALL internal whitespace for matching: provincial headers often
        # space out CJK chars for alignment ("项 目 名 称"), which would defeat a
        # substring match against "项目名称".
        cells = [re.sub(r"\s+", "", _norm(c)) for c in rows[ri]]
        mapping: dict[str, int] = {}
        for role, kws in HEADER_KEYWORDS:
            if role in mapping:
                continue
            for ci, cell in enumerate(cells):
                if ci in mapping.values():
                    continue
                if any(kw in cell for kw in kws):
                    mapping[role] = ci
                    break
        # A real header must at least identify the project-title column.
        if "display_name" in mapping:
            score = len(mapping)
            if best is None or score > best[2]:
                best = (ri, mapping, score)
    if best is None:
        return None
    return best[0], best[1]


def rows_to_records(rows: list[list[Any]], cfg: ProvinceConfig, source_file: str,
                    landing_page: str, article_title: str) -> list[dict]:
    """Apply header detection + row extraction to one table's rows."""
    hdr = detect_header(rows)
    if hdr is None:
        return []
    hi, mapping = hdr
    title_ci = mapping["display_name"]
    scheme_default = cfg.scheme_from_title(article_title) if cfg.scheme_from_title else None
    year_default = cfg.year_from_title(article_title) if cfg.year_from_title else None
    out = []
    current_scheme = scheme_default
    for row in rows[hi + 1:]:
        cells = [_norm(c) for c in row]
        title = cells[title_ci] if title_ci < len(cells) else ""
        pi = cells[mapping["lead_name"]] if mapping.get("lead_name", 99) < len(cells) else ""
        inst = cells[mapping["institution"]] if mapping.get("institution", 99) < len(cells) else ""
        # Section-heading rows (e.g. "1.区块链安全...") carry a title but no PI/inst;
        # capture them as the running scheme/direction and skip.
        if title and not pi and not inst and SECTION_ROW.match(title):
            current_scheme = title
            continue
        if not title:
            continue
        # A funded-project row needs at least a title and (PI or institution).
        if not pi and not inst:
            continue
        # PI name normalization: split off a "/" leader==team form ("郝京诚/郝京诚").
        if "/" in pi:
            pi = pi.split("/", 1)[0].strip()
        scheme = ""
        if mapping.get("funder_scheme", 99) < len(cells):
            scheme = cells[mapping["funder_scheme"]]
        scheme = scheme or (current_scheme if current_scheme != scheme_default else "") or scheme_default or ""
        award_id = ""
        if mapping.get("funder_award_id", 99) < len(cells):
            award_id = cells[mapping["funder_award_id"]]
        amount = ""
        if mapping.get("amount", 99) < len(cells):
            amount = cells[mapping["amount"]]
        year = ""
        if mapping.get("start_year", 99) < len(cells):
            year = cells[mapping["start_year"]]
        if not year and year_default:
            year = str(year_default)
        # Chinese PI: full name -> family_name, given_name NULL (NSFC precedent).
        fam = pi
        giv = None
        out.append({
            "funder_award_id": award_id or None,
            "display_name": title,
            "funder_scheme": scheme or None,
            "lead_family_name": fam or None,
            "lead_given_name": giv,
            "institution": inst or None,
            "amount_raw": amount or None,
            "start_year": year or None,
            "landing_page_url": landing_page,
            "source_file": source_file,
            "article_title": article_title,
            "provenance": cfg.provenance,
        })
    return out


# ---------------------------------------------------------------------------
# Attachment parsers -> list[list[cell]]
# ---------------------------------------------------------------------------

def _sniff(path: Path) -> str:
    head = path.read_bytes()[:8]
    if head.startswith(b"%PDF"):
        return "pdf"
    if head.startswith(b"PK\x03\x04") or head.startswith(b"PK"):
        try:
            names = zipfile.ZipFile(path).namelist()
        except zipfile.BadZipFile:
            return "?"
        if any(n.startswith("xl/") for n in names):
            return "xlsx"
        if any(n.startswith("word/") for n in names):
            return "docx"
        return "zip"
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return "ole"  # legacy .doc or BIFF .xls
    return "?"


def parse_xlsx(path: Path) -> list[list[list[Any]]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(_norm(c) for c in r)]
        if len(rows) >= 2:
            tables.append(rows)
    return tables


def parse_xls_biff(path: Path) -> list[list[list[Any]]]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    tables = []
    for sh in wb.sheets():
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        rows = [r for r in rows if any(_norm(c) for c in r)]
        if len(rows) >= 2:
            tables.append(rows)
    return tables


def parse_docx(path: Path) -> list[list[list[Any]]]:
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8", "replace")
    tables = []
    for tbl in re.findall(r"<w:tbl[ >].*?</w:tbl>", xml, re.S):
        rows = []
        for tr in re.findall(r"<w:tr[ >].*?</w:tr>", tbl, re.S):
            cells = []
            for tc in re.findall(r"<w:tc[ >].*?</w:tc>", tr, re.S):
                cells.append("".join(re.findall(r"<w:t[^>]*>([^<]*)</w:t>", tc)))
            rows.append(cells)
        rows = [r for r in rows if any(_norm(c) for c in r)]
        if len(rows) >= 2:
            tables.append(rows)
    return tables


def parse_pdf(path: Path) -> list[list[list[Any]]]:
    import pdfplumber
    tables = []
    with pdfplumber.open(path) as pdf:
        carry_header = None
        for page in pdf.pages:
            found = page.extract_tables()
            if not found:
                # Fall back to text-positioning strategy (older SD PDFs have no ruled lines).
                found = page.extract_tables(
                    {"vertical_strategy": "text", "horizontal_strategy": "text"})
            for t in found:
                t = [[_norm(c) for c in row] for row in t if any(_norm(c) for c in row)]
                if not t:
                    continue
                # Continuation pages repeat no header; reattach the last seen header
                # so detect_header still fires. We flag by prepending a synthetic
                # header only if this page's first row lacks title keywords.
                if carry_header and not any("项目" in c or "名称" in c or "课题" in c for c in t[0]):
                    t = [carry_header] + t
                else:
                    carry_header = t[0]
                tables.append(t)
    return tables


PARSERS: dict[str, Callable[[Path], list[list[list[Any]]]]] = {
    "pdf": parse_pdf,
    "xlsx": parse_xlsx,
    "docx": parse_docx,
}


def parse_attachment(path: Path) -> list[list[list[Any]]]:
    kind = _sniff(path)
    if kind == "ole":
        # Could be a BIFF .xls (openable by xlrd) or a legacy .doc (not handled).
        try:
            return parse_xls_biff(path)
        except Exception:
            print(f"    [WARN] legacy OLE .doc not parseable, skipping: {path.name}")
            return []
    fn = PARSERS.get(kind)
    if fn is None:
        print(f"    [WARN] unknown attachment type ({kind}): {path.name}")
        return []
    try:
        return fn(path)
    except Exception as exc:
        print(f"    [WARN] parse failed ({kind}) for {path.name}: {exc}")
        return []


# ---------------------------------------------------------------------------
# HTTP + listing enumeration
# ---------------------------------------------------------------------------

def make_session(cfg: ProvinceConfig) -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": UA})
    s.verify = False
    return s


def enumerate_listing(cfg: ProvinceConfig, col: ListingColumn, sess: requests.Session,
                      max_articles: Optional[int] = None) -> list[dict]:
    """Page through one Hanweb column; return [{url,title,date}] (deduped)."""
    if col.needs_cookie_prime:
        try:
            sess.get(col.referer, timeout=30)
        except Exception as exc:
            print(f"  [WARN] cookie prime failed: {exc}")
    proxy = col.proxy_url(cfg.base_url)
    headers = {"Referer": col.referer, "X-Requested-With": "XMLHttpRequest"}
    records: list[dict] = []
    seen: set[str] = set()
    total: Optional[int] = None
    start = 1
    consec_err = 0
    while True:
        end = start + col.stride - 1
        url = f"{proxy}?startrecord={start}&endrecord={end}&perpage={col.perpage}"
        try:
            r = sess.post(url, data=col.post_data(), headers=headers, timeout=30)
        except Exception as exc:
            consec_err += 1
            print(f"  req {start}-{end}: EXC {exc} ({consec_err}/5)")
            if consec_err >= 5:
                raise
            time.sleep(5)
            continue
        if r.status_code != 200 or "<datastore>" not in r.text:
            consec_err += 1
            print(f"  req {start}-{end}: HTTP {r.status_code} len {len(r.text)} ({consec_err}/5)")
            if consec_err >= 5:
                raise RuntimeError("too many consecutive listing failures")
            time.sleep(5)
            continue
        consec_err = 0
        if total is None:
            m = re.search(r"<totalrecord>(\d+)</totalrecord>", r.text)
            total = int(m.group(1)) if m else None
            print(f"  column {col.columnid}: totalrecord={total}")
        for rec in re.findall(r"<record><!\[CDATA\[(.*?)\]\]></record>", r.text, re.S):
            am = (re.search(r'href="([^"]+)"[^>]*title="([^"]*)"', rec)
                  or re.search(r'href="([^"]+)"[^>]*>(.*?)</a>', rec, re.S))
            if not am:
                continue
            href = am.group(1)
            if href in seen:
                continue
            seen.add(href)
            title = re.sub(r"<[^>]+>", "", am.group(2)).strip()
            dm = re.search(r"(\d{4}-\d{2}-\d{2})", rec)
            records.append({"url": href, "title": title, "date": dm.group(1) if dm else None})
        print(f"  fetched {start}-{end}: cum {len(records)}")
        if max_articles and len(records) >= max_articles:
            break
        start = end + 1
        if total and start > total:
            break
        time.sleep(cfg.request_pause)
    return records


def select_award_articles(cfg: ProvinceConfig, records: list[dict]) -> list[dict]:
    inc = re.compile(cfg.article_pattern)
    exc = re.compile(cfg.article_exclude)
    return [r for r in records if inc.search(r["title"]) and not exc.search(r["title"])]


# ---------------------------------------------------------------------------
# Article + attachment processing
# ---------------------------------------------------------------------------

def article_url(cfg: ProvinceConfig, href: str) -> str:
    return href if href.startswith("http") else urljoin(cfg.base_url, href)


def find_attachment_links(cfg: ProvinceConfig, html: str) -> list[tuple[str, str]]:
    """Return [(href, link_text)] for downloadable attachments in an article."""
    body = re.sub(r"<script.*?</script>", "", html, flags=re.S)
    pat = re.compile(cfg.attachment_ext, re.I)
    out = []
    for m in re.finditer(r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', body, re.S):
        href, text = m.group(1), re.sub(r"<[^>]+>", "", m.group(2)).strip()
        if pat.search(href):
            out.append((href, text))
    return out


def extract_inline_table(html: str) -> list[list[list[Any]]]:
    """Parse the article's own <table> when there is no attachment."""
    body = re.sub(r"<script.*?</script>", "", html, flags=re.S)
    tables = []
    for tbl in re.findall(r"<table[ >].*?</table>", body, re.S):
        rows = []
        for tr in re.findall(r"<tr[ >].*?</tr>", tbl, re.S):
            cells = [re.sub(r"<[^>]+>", "", c).strip()
                     for c in re.findall(r"<t[dh][ >].*?</t[dh]>", tr, re.S)]
            rows.append(cells)
        rows = [r for r in rows if any(_norm(c) for c in r)]
        if len(rows) >= 2:
            tables.append(rows)
    return tables


def process_article(cfg: ProvinceConfig, art: dict, sess: requests.Session,
                    work_dir: Path) -> list[dict]:
    url = article_url(cfg, art["url"])
    try:
        r = sess.get(url, timeout=45)
        r.raise_for_status()
    except Exception as exc:
        print(f"  [WARN] article fetch failed {url}: {exc}")
        return []
    html = r.content.decode("utf-8", "replace")
    records: list[dict] = []
    links = find_attachment_links(cfg, html)
    adir = work_dir / "attachments"
    adir.mkdir(parents=True, exist_ok=True)
    for href, text in links:
        full = href if href.startswith("http") else urljoin(cfg.base_url, href)
        # Derive a stable local filename.
        tail = re.sub(r"[^\w.]", "_", (full.split("filename=", 1)[-1] if "downfile" in full else urlparse(full).path.rsplit("/", 1)[-1]))
        ext = ""
        m = re.search(r"\.(xlsx?|pdf|docx?|et|wps)", (text + " " + tail), re.I)
        if m:
            ext = "." + m.group(1).lower()
        local = adir / f"{art['url'].rsplit('/',1)[-1]}__{tail[:50]}{ext}"
        if not local.exists():
            try:
                fr = sess.get(full, timeout=90)
                local.write_bytes(fr.content)
                time.sleep(cfg.request_pause)
            except Exception as exc:
                print(f"    [WARN] attachment download failed {full}: {exc}")
                continue
        tables = parse_attachment(local)
        for t in tables:
            records += rows_to_records(t, cfg, source_file=text or local.name,
                                       landing_page=url, article_title=art["title"])
    if not links and cfg.allow_inline_table:
        for t in extract_inline_table(html):
            records += rows_to_records(t, cfg, source_file="(inline table)",
                                       landing_page=url, article_title=art["title"])
    print(f"  {art.get('date','?')} | {art['title'][:44]} | {len(links)} attach -> {len(records)} rows")
    return records


# ---------------------------------------------------------------------------
# Checkpointing + output + upload
# ---------------------------------------------------------------------------

def s3_key(provenance: str) -> str:
    return f"awards/{provenance}/{provenance}_projects.parquet"


def check_no_shrink(provenance: str, new_count: int, allow_shrink: bool, work_dir: Path) -> bool:
    """Runbook 1.4: refuse to overwrite S3 with a smaller corpus."""
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError as exc:
        raise RuntimeError("boto3 required for shrink-check; use --skip-upload for local runs") from exc
    key = s3_key(provenance)
    client = boto3.client("s3")
    print(f"  Re-ingest safety check vs s3://{S3_BUCKET}/{key}")
    try:
        client.head_object(Bucket=S3_BUCKET, Key=key)
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in ("404", "NoSuchKey", "NotFound"):
            print("    no existing parquet: first ingest, no shrink check needed.")
            return True
        print(f"    [WARN] head_object failed ({code}); treating as first ingest.")
        return True
    prev = work_dir / f"_prev_{provenance}.parquet"
    try:
        client.download_file(S3_BUCKET, key, str(prev))
        prev_count = len(pd.read_parquet(prev))
    except Exception as exc:
        print(f"    [ERROR] could not read existing parquet ({exc}); aborting upload.")
        return False
    finally:
        prev.unlink(missing_ok=True)
    print(f"    previous count: {prev_count:,}   new count: {new_count:,}")
    if new_count < prev_count and not allow_shrink:
        print(f"\n[ERROR] Refusing to shrink {provenance} corpus ({prev_count:,} -> {new_count:,}).")
        return False
    if new_count < prev_count:
        print("    [OVERRIDE] --allow-shrink set.")
    else:
        print("    [OK] new corpus is not smaller; safe to overwrite.")
    return True


def write_and_upload(cfg: ProvinceConfig, df: pd.DataFrame, output_dir: Path,
                     skip_upload: bool, allow_shrink: bool) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    parquet_path = output_dir / f"{cfg.provenance}_projects.parquet"
    df = df.astype("string")  # force string dtype before to_parquet (runbook §1.2.5)
    df.to_parquet(parquet_path, index=False, engine="pyarrow")
    size_mb = parquet_path.stat().st_size / (1024 * 1024)
    print(f"\n[OK] wrote {len(df):,} rows ({size_mb:.2f} MB) -> {parquet_path}")
    # Coverage summary (helps the §6.7 amount-coverage decision).
    if len(df):
        amt = df["amount_raw"].notna().mean() * 100 if "amount_raw" in df else 0
        pi = df["lead_family_name"].notna().mean() * 100
        inst = df["institution"].notna().mean() * 100
        print(f"     coverage: PI {pi:.1f}%  institution {inst:.1f}%  amount {amt:.1f}%")
        years = pd.to_numeric(df["start_year"], errors="coerce").dropna()
        if len(years):
            print(f"     start_year range: {int(years.min())}-{int(years.max())}")
    if skip_upload:
        print("[SKIP] --skip-upload set; not uploading to S3.")
        return
    if not check_no_shrink(cfg.provenance, len(df), allow_shrink, output_dir):
        raise SystemExit(1)
    uri = f"s3://{S3_BUCKET}/{s3_key(cfg.provenance)}"
    print(f"  Uploading -> {uri}")
    subprocess.run(["aws", "s3", "cp", str(parquet_path), uri], check=True)
    print(f"  [OK] uploaded to {uri}")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_province(cfg: ProvinceConfig, argv: Optional[list[str]] = None) -> None:
    parser = argparse.ArgumentParser(description=f"{cfg.funder_display_name} awards -> S3")
    parser.add_argument("--output-dir", type=Path, default=Path(f"/tmp/{cfg.provenance}"))
    parser.add_argument("--limit", type=int, default=None,
                        help="Max award-articles to process (smoke testing)")
    parser.add_argument("--skip-upload", action="store_true")
    parser.add_argument("--allow-shrink", action="store_true")
    parser.add_argument("--skip-listing", action="store_true",
                        help="Reuse cached listing.json / articles from output-dir")
    args = parser.parse_args(argv)

    wd = args.output_dir
    wd.mkdir(parents=True, exist_ok=True)
    print("=" * 64)
    print(f"{cfg.funder_display_name}  (F{cfg.funder_id})")
    print(f"  provenance: {cfg.provenance}")
    print(f"  base:       {cfg.base_url}")
    print(f"  started:    {datetime.now(timezone.utc).isoformat()}")
    print("=" * 64)

    sess = make_session(cfg)

    # 1. Listing
    listing_cache = wd / "listing.json"
    if args.skip_listing and listing_cache.exists():
        all_records = json.loads(listing_cache.read_text())
        print(f"[cache] {len(all_records)} listing records")
    else:
        all_records = []
        for col in cfg.listing_columns:
            all_records += enumerate_listing(cfg, col, sess)
        # dedup by url
        seen, dedup = set(), []
        for r in all_records:
            if r["url"] not in seen:
                seen.add(r["url"])
                dedup.append(r)
        all_records = dedup
        listing_cache.write_text(json.dumps(all_records, ensure_ascii=False, indent=1))
        print(f"[OK] {len(all_records)} listing records -> {listing_cache}")

    # 2. Filter to award lists
    awards = select_award_articles(cfg, all_records)
    print(f"[OK] {len(awards)} award-list articles match config pattern")
    if args.limit:
        awards = awards[:args.limit]
        print(f"    (--limit {args.limit})")

    # 3. Process with checkpointing
    ckpt = wd / "records.checkpoint.jsonl"
    done_urls: set[str] = set()
    if ckpt.exists():
        for line in ckpt.read_text().splitlines():
            if line.strip():
                done_urls.add(json.loads(line)["_article_url"])
        print(f"[resume] {len(done_urls)} articles already checkpointed")
    with open(ckpt, "a", encoding="utf-8") as fh:
        for art in awards:
            if art["url"] in done_urls:
                continue
            recs = process_article(cfg, art, sess, wd)
            for rec in recs:
                rec["_article_url"] = art["url"]
                fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
            if not recs:
                # Still checkpoint the article as processed (0 rows) so resume skips it.
                fh.write(json.dumps({"_article_url": art["url"], "_empty": True},
                                    ensure_ascii=False) + "\n")
            fh.flush()

    # 4. Assemble dataframe
    rows = []
    for line in ckpt.read_text().splitlines():
        if not line.strip():
            continue
        obj = json.loads(line)
        if obj.get("_empty"):
            continue
        obj.pop("_article_url", None)
        rows.append(obj)
    if not rows:
        print("[ERROR] no records extracted; check config pattern / parsing.")
        raise SystemExit(1)
    df = pd.DataFrame(rows).drop_duplicates(
        subset=["display_name", "lead_family_name", "institution", "provenance"])
    print(f"[OK] {len(df):,} unique project rows after dedup")

    write_and_upload(cfg, df, args.output_dir, args.skip_upload, args.allow_shrink)
